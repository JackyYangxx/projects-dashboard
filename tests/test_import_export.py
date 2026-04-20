"""
E2E tests for Dashboard import/export functionality.
Usage:
  python3 /Users/fxy/.claude/skills/webapp-testing/scripts/with_server.py \
    --server "npm run dev" --port 5173 \
    -- python3 tests/test_import_export.py
"""

import os
import re
import tempfile
from datetime import date
from playwright.sync_api import sync_playwright
import openpyxl

BASE_URL = "http://localhost:5173"
REQUIRED_HEADERS = ["项目名称", "产品线", "负责人", "状态", "项目进展", "总预算", "已用预算"]


def make_xlsx(rows: list[dict], path: str):
    """Create an xlsx file from a list of dicts (headers = keys of first row)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
    wb.save(path)


def valid_row(name="测试项目A", product="测试产品线", leader="张三") -> dict:
    return {
        "项目名称": name,
        "产品线": product,
        "负责人": leader,
        "状态": "进行中",
        "项目进展": 50,
        "总预算": 1000000,
        "已用预算": 500000,
    }


def wait_for_dashboard(page):
    page.goto(BASE_URL)
    page.wait_for_load_state("networkidle")
    # Ensure the import/export buttons are visible
    page.wait_for_selector("text=导入", timeout=10000)


def test_export(page):
    """Export triggers a download with correct filename format."""
    print("\n[TEST] Export — should download projects_YYYY-MM-DD.xlsx")
    wait_for_dashboard(page)

    today = date.today().isoformat()
    expected_filename = f"projects_{today}.xlsx"

    with page.expect_download(timeout=10000) as dl_info:
        page.get_by_text("导出").click()

    download = dl_info.value
    assert download.suggested_filename == expected_filename, (
        f"Expected '{expected_filename}', got '{download.suggested_filename}'"
    )

    # Save and verify it's a valid xlsx
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        download.save_as(tmp.name)
        wb = openpyxl.load_workbook(tmp.name)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        os.unlink(tmp.name)

    assert "项目名称" in headers, f"Export missing '项目名称', got: {headers}"
    assert "负责人" in headers, f"Export missing '负责人'"
    print(f"  PASS — downloaded '{expected_filename}', headers: {headers}")


def test_import_valid(page):
    """Import valid xlsx inserts rows and shows success alert."""
    print("\n[TEST] Import valid data — should show 导入完成 alert")
    wait_for_dashboard(page)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = tmp.name

    try:
        make_xlsx([
            valid_row("E2E项目Alpha", "Alpha线", "王五"),
            valid_row("E2E项目Beta", "Beta线", "李四"),
        ], path)

        dialog_messages = []
        page.on("dialog", lambda d: (dialog_messages.append(d.message), d.accept()))

        with page.expect_file_chooser(timeout=5000) as fc_info:
            page.get_by_text("导入").click()

        fc_info.value.set_files(path)
        page.wait_for_timeout(1500)  # wait for FileReader + alert

        assert len(dialog_messages) == 1, f"Expected 1 dialog, got: {dialog_messages}"
        msg = dialog_messages[0]
        assert "导入完成" in msg, f"Expected '导入完成' in alert, got: '{msg}'"
        assert "成功 2 条" in msg, f"Expected '成功 2 条', got: '{msg}'"
        print(f"  PASS — alert: '{msg}'")
    finally:
        os.unlink(path)


def test_import_missing_headers(page):
    """Import xlsx with missing required columns shows error alert."""
    print("\n[TEST] Import missing headers — should show 缺少必要字段 alert")
    wait_for_dashboard(page)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = tmp.name

    try:
        make_xlsx([{"项目名称": "仅名称项目"}], path)

        dialog_messages = []
        page.on("dialog", lambda d: (dialog_messages.append(d.message), d.accept()))

        with page.expect_file_chooser(timeout=5000) as fc_info:
            page.get_by_text("导入").click()

        fc_info.value.set_files(path)
        page.wait_for_timeout(1500)

        assert len(dialog_messages) == 1, f"Expected 1 dialog, got: {dialog_messages}"
        msg = dialog_messages[0]
        assert "缺少必要字段" in msg, f"Expected '缺少必要字段' in alert, got: '{msg}'"
        print(f"  PASS — alert: '{msg}'")
    finally:
        os.unlink(path)


def test_import_empty_file(page):
    """Import empty xlsx shows error alert."""
    print("\n[TEST] Import empty file — should show Excel 文件为空 alert")
    wait_for_dashboard(page)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = tmp.name

    try:
        make_xlsx([], path)

        dialog_messages = []
        page.on("dialog", lambda d: (dialog_messages.append(d.message), d.accept()))

        with page.expect_file_chooser(timeout=5000) as fc_info:
            page.get_by_text("导入").click()

        fc_info.value.set_files(path)
        page.wait_for_timeout(1500)

        assert len(dialog_messages) == 1, f"Expected 1 dialog, got: {dialog_messages}"
        msg = dialog_messages[0]
        assert "文件为空" in msg, f"Expected '文件为空' in alert, got: '{msg}'"
        print(f"  PASS — alert: '{msg}'")
    finally:
        os.unlink(path)


def test_import_skip_invalid_rows(page):
    """Rows with empty name/leader or invalid status are skipped."""
    print("\n[TEST] Import invalid rows — should skip and report correctly")
    wait_for_dashboard(page)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = tmp.name

    try:
        make_xlsx([
            valid_row("有效项目", "线A", "赵六"),          # valid → success
            {**valid_row(), "项目名称": ""},               # empty name → skip
            {**valid_row(), "负责人": ""},                 # empty leader → skip
            {**valid_row("X"), "状态": "未知状态"},        # bad status → skip
        ], path)

        dialog_messages = []
        page.on("dialog", lambda d: (dialog_messages.append(d.message), d.accept()))

        with page.expect_file_chooser(timeout=5000) as fc_info:
            page.get_by_text("导入").click()

        fc_info.value.set_files(path)
        page.wait_for_timeout(1500)

        assert len(dialog_messages) == 1, f"Expected 1 dialog, got: {dialog_messages}"
        msg = dialog_messages[0]
        assert "成功 1 条" in msg, f"Expected '成功 1 条', got: '{msg}'"
        assert "跳过 3 条" in msg, f"Expected '跳过 3 条', got: '{msg}'"
        print(f"  PASS — alert: '{msg}'")
    finally:
        os.unlink(path)


def main():
    tests = [
        test_export,
        test_import_valid,
        test_import_missing_headers,
        test_import_empty_file,
        test_import_skip_invalid_rows,
    ]

    passed = 0
    failed = 0

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        for test_fn in tests:
            context = browser.new_context(accept_downloads=True)
            page = context.new_page()
            try:
                test_fn(page)
                passed += 1
            except Exception as e:
                print(f"  FAIL — {e}")
                failed += 1
            finally:
                context.close()
        browser.close()

    print(f"\n{'='*40}")
    print(f"Results: {passed} passed, {failed} failed")
    if failed:
        exit(1)


if __name__ == "__main__":
    main()
